# -*- coding: utf-8 -*-
"""AI

Automatically generated by Colaboratory.

Original file is located at
    https://colab.research.google.com/drive/1Nlfg6lmnSg1s3OuWiNYh_Ek30zzYUPFM
"""

print(" welcome to python!")

print("welcome", "to","python")

name = input("what's your name?")

value1 = input('enter first number:')

value2= input('enter second number:')

value1 + value2

value = input('enter an integer:')
value = int (value)

another_value = int(input('enter another integer:'))

value+another_value

with open('test.txt',mode='w') as accounts: 
 accounts.write('100 Jone 24.98\n')
 accounts.write('200 Doe 345.67\n')
 accounts.write('300 White 0.00\n')
 accounts.write('400 Store 42.16\n')
 accounts.write('500 Rich 224.62\n')

with open('test1.txt',mode='w') as accounts: 
 print('100 Jone 24.98', file = accounts)
 print('200 Doe 345.67', file=accounts)
 print('300 White 0.00', file=accounts)
 print('400 Store 42.16', file=accounts)
 print('500 Rich 224.62', file=accounts)

from google.colab import files

uploaded = files.upload()

with open('test1.txt',mode='w') as accounts: 
 print('100 Jone 24.98', file = accounts)
 print('200 Doe 345.67', file=accounts)
 print('300 White 0.00', file=accounts)
 print('400 Store 42.16', file=accounts)
 print('500 Rich 224.62', file=accounts)

"""# Mục mới"""

with open('test1.txt', mode='r') as accounts:
  print(f'{"Account":<10}{"name":<10}{"Balance":>10}')
  for record in accounts:
    Account, name, Balance = record.split()
    print(f'{account:<10}{name:<10}{Balance:>10}')

from google.colab import files
uploaded = files.upload( )

#load library
import pandas as pd
# Creat URL
url = 'file.csv'
# load data
dataframe = pd.read_csv(url)
#view first two rows
dataframe.head(5)

from google.colab import files
uploaded = files.upload()

#load library
import pandas as pd
# Creat URL
url = 'CSV.csv'
# load data
dataframe = pd.read_csv(url)
#select first rows
dataframe.iloc[2]

#load library
import pandas as pd
# Creat URL
url = 'CSV.csv'
# load data
dataframe = pd.read_csv(url)
#select first rows
dataframe.iloc[0:4]

from google.colab import files
uploaded = files.upload()

#set index
dataframe = dataframe.set_index(dataframe['FULL NAME'])
#show row
dataframe.loc['BACH']

#load library
import pandas as pd
#create url
url ='CSV.csv'
#load data
dataframe = pd.read_csv(url)
#view first two rows
dataframe.head(3)

#load library
import pandas as pd
#create url
url ='CSV.csv'
#load data
dataframe = pd.read_csv(url)
#view first two rows
dataframe.head(9)

from google.colab import files
uploaded = files.upload()

#load library
import pandas as pd
#create url
url ='loc.csv'
#load data
dataframe = pd.read_csv(url)
#view first two rows
dataframe.head(5)

#filter rows 
dataframe[(dataframe['faculty']=='ckm' ) & (dataframe['year']==18)]

from google.colab import files
uploaded = files.upload()
#load library
import pandas as pd
#create url
url ='loc.csv'
#load data
dataframe = pd.read_csv(url)
#Rename column,show 2 rows
dataframe.rename(columns= {'year':'SCORE'}).head(2)

from google.colab import files
uploaded = files.upload()

from google.colab import files
uploaded = files.upload()

#load library
import pandas as pd
#create url
url ='loc.csv'
#load data
dataframe = pd.read_csv(url)
#delete column xoa cot faculty
dataframe.drop('faculty', axis=1).head(4)

#load library
import pandas as pd
#create url
url ='loc.csv'
#load data
dataframe = pd.read_csv(url)
#delete rows xoa hang co diem = 11
dataframe[dataframe['SCORE']!= 11].head(4)

#drop duplicates,show first two rows of output xoa hang trung
dataframe.drop_duplicates().head(5)

#load library
import pandas as pd
#create url
url ='loc.csv'
#load data
dataframe = pd.read_csv(url)
# gop hang theo gia tri year
dataframe.groupby('year').mean()

from google.colab import files
uploaded = files.upload()

#load library
import pandas as pd
# Creat URL
url = 'BT.csv'
# load data
dataframe = pd.read_csv(url)
#view first two rows
dataframe.head(10)

from google.colab import files
uploaded = files.upload()

#load library
import pandas as pd
#create url
url ='BT.csv'
#load data
dataframe = pd.read_csv(url)
# gop hang theo gia tri year
dataframe.groupby('Event').mean()

#load library
import pandas as pd
#create url
url ='BT.csv'
#load data
dataframe = pd.read_csv(url)
for name in dataframe['City'][0:30]: print(name.upper())

import pandas as pd
#create url
url ='BT.csv'
#load data
dataframe = pd.read_csv(url)
dataframe['NOC'].unique()

from google.colab import files
uploaded = files.upload()
import pandas as pd
#create url
url ='BT.csv'
#load data
dataframe= pd.read_csv(url)
dataframe['Year'].unique()

import pandas as pd
#create url
url ='BT.csv'
#load data
data= pd.read_csv(url)
data.groupby(['Year','Medal']).count()

import pandas as pd
#create url
url ='BT.csv'
#load data
data= pd.read_csv(url)
data.groupby(['Year','NOC'])['Medal'].count()

from openpyxl import Workbook
wb=Workbook()
sheet = wb.active

sheet['A1'] ='hello'
sheet['B1'] ='world'
sheet['A2'] = 'Bach'
sheet['B2'] = '012345'
wb.save('test.xlsx')

import openpyxl
wb=openpyxl.load_workbook('test.xlsx')
sheet=wb.active
t1=sheet['A1']
t2=sheet['B1']
t3=sheet.cell(row = 2,column=1)
t4=sheet.cell(row = 2,column=2)
print(t1.value,t2.value,t3.value,t4.value)

import openpyxl
wb = openpyxl.Workbook()
sheet  = wb.active
c1=sheet.cell(row=1,column=1)
c1.value='Nguyen Van A'
c2 = sheet.cell(row=1,column=2)
c2.value='19146001'
c3 = sheet['A2']
c3.value='BACHPRO'
c4=sheet['B2']
c4.value='19146002'
wb.save('test1.xlsx')

import openpyxl 
wb=openpyxl.Workbook()
sheet=wb.active
wb.create_sheet(index=1,title='0379495967')
sheet=wb.active
wb.create_sheet(index=2,title='0966949541')
wb.save('test2.xlsx')

import datetime
thoigian=datetime.datetime.now()
print(thoigian)
#in gio

from datetime import date
today=date.today()
print('nam',today.year)
print('thang',today.month)
print('ngay',today.day)

import time 
time = time.localtime(time.time())
print('thoi gian hien tai',time) #tm_day:ngayf thu 66 cua nam #tm_yday= 0(ngay thu 2)

import time
time  = time.asctime(time.localtime(time.time()))
print('Thoi gian hien tai:',time)

import calendar
cal = calendar.month(2001,9)
print('day la thang sinh cua BACH')
print(cal)

calendar.calendar(2001,w=2,l=1,c=6) # w chieu rong cua moi ki tu__ L so dong cua moi

calendar.monthcalendar(2022,3)

calendar.isleap(2024)

calendar.leapdays(1,2020)

for(i=1,i<=2020, )
print('nam nhuan')

pip install scikit-fuzzy



import numpy as np



import skfuzzy as fuzz



#tao khoang mo
x = np.arange(11)
x

#tao ham lien thuoc
mfx = fuzz.trimf(x,[0,5,10])
mfx

import numpy as np
import skfuzzy as fuzz
import matplotlib.pyplot as plt
#tao khoang mo
x= np.arange(30,81,1)
#tao ham lien thuoc
slow=fuzz.trimf(x,[0,30,50])
medium=fuzz.trimf(x,[30,50,70])
mediumfast=fuzz.trimf(x,[50,60,70])
fullspeed=fuzz.trimf(x,[60,80,100])



import numpy as np
import skfuzzy as fuzz
import matplotlib.pyplot as plt

#tao khoang mo
x= np.arange(30,81,1)
#tao ham lien thuoc
slow = fuzz.trimf(x,[0,30,50])
medium = fuzz.trimf(x,[30,50,60])
mediumfast = fuzz.trimf(x,[50,60,80])
fullspeed = fuzz.trimf(x,[70,80,100])
#ve ham lien thuoc
plt.figure()
plt.plot(x,slow,'b',linewidth = 1.5,label =' slow')
plt.plot(x,medium,'p',linewidth = 1.5,label =' medium')
plt.plot(x,mediumfast,'r',linewidth = 1.5,label =' mediumfast')
plt.plot(x,fullspeed,'y',linewidth = 1.5,label =' fullspeed')

plt.title('Bieu do toc do')



plt.xlabel('khoang gia tri ')



plt.ylabel('gia tri ham lien thuoc')

plt.legend(loc='center right', ) #@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@2

#tao khoang mo
x= np.arange(30,81,1)
#tao ham lien thuoc
slow = fuzz.trapmf(x,[30,30,30,50])
medium = fuzz.trapmf(x,[30,50,50,70])
mediumfast = fuzz.trapmf(x,[50,60,60,80])
fullspeed = fuzz.trapmf(x,[60,80,80,80])
#ve ham lien thuoc
plt.figure()
plt.plot(x,slow,'b',linewidth = 1.5,label =' slow')
plt.plot(x,medium,'k',linewidth = 1.5,label =' medium')
plt.plot(x,mediumfast,'r',linewidth = 1.5,label =' mediumfast')
plt.plot(x,fullspeed,'y',linewidth = 1.5,label =' fullspeed')
#slow=fuzz.trapmf(x,[30,30,30,50])
#medium=fuzz.trapmf(x,[30,50,50,70])
#medium_fast=fuzz.trapmf(x,[50,60,60,80])
#full_speed=fuzz.trapmf(x,[60,80,80,80])

#tao khoang mo
x= np.arange(30,81,1)
#tao ham lien thuoc
slow = fuzz.gaussmf(x,30,10)
medium = fuzz.gaussmf(x,40,15)
mediumfast = fuzz.gaussmf(x,50,10)
fullspeed = fuzz.gaussmf(x,60,10)
#ve ham lien thuoc
plt.figure()
plt.plot(x,slow,'b',linewidth = 1.5,label =' slow')
plt.plot(x,medium,'k',linewidth = 1.5,label =' medium')
plt.plot(x,mediumfast,'r',linewidth = 1.5,label =' mediumfast')
plt.plot(x,fullspeed,'y',linewidth = 1.5,label =' fullspeed')



#tao khoang mo
x= np.arange(30,81,1)
#tao ham lien thuoc
slow = fuzz.gbellmf(x,10,20,50)
medium = fuzz.gbellmf(x,10,20,40)
mediumfast = fuzz.gbellmf(x,40,10,40)
fullspeed = fuzz.gbellmf(x,50,20,30)
#ve ham lien thuoc
plt.figure()
plt.plot(x,slow,'b',linewidth = 1.5,label =' slow')
plt.plot(x,medium,'k',linewidth = 1.5,label =' medium')
plt.plot(x,mediumfast,'r',linewidth = 1.5,label =' mediumfast')
plt.plot(x,fullspeed,'y',linewidth = 1.5,label =' fullspeed')
#x,[a,b,c]

import numpy as np
import skfuzzy as fuzz

from skfuzzy import control as ctrl # thu vien ctrl

# tao tin hieu vao va ra
food=ctrl.Antecedent(np.arange(0,11,1),'food') 
service = ctrl.Antecedent(np.arange(0,11,1),'service')

tip = ctrl.Consequent(np.arange(0,11,1),'tip')

#tao ham lien thuoc
food.automf(3)
food.view()
service.automf(3)
service.view()

#tao ham lien thuoc bang tay
tip['less']=fuzz.trimf(tip.universe,[0,0,5])
tip['medium']=fuzz.trimf(tip.universe,[0,5,10])
tip['much']=fuzz.trimf(tip.universe,[5,10,10])
tip.view()

rule1 = ctrl.Rule(food['poor'] | service['poor'] ,tip['less'])
rule2 = ctrl.Rule(food['poor'] | service['average'] ,tip['less'])
rule3 = ctrl.Rule(food['poor'] | service['good'] ,tip['medium'])
rule4 = ctrl.Rule(food['average'] | service['poor'] ,tip['less'])
rule5 = ctrl.Rule(food['average'] | service['average'] ,tip['medium'])
rule6 = ctrl.Rule(food['average'] | service['good'] ,tip['much'])
rule7 = ctrl.Rule(food['good'] | service['poor'] ,tip['medium'])
rule8 = ctrl.Rule(food['good'] | service['average'] ,tip['medium'])
rule9 = ctrl.Rule(food['good'] | service['good'] ,tip['much'])

#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@

tipping_ctrl=ctrl.ControlSystem([rule1,rule2,rule3])
tipping = ctrl.ControlSystemSimulation(tipping_ctrl)

tipping.input['food']=7.5
tipping.input['service']=6
tipping.compute()
print(tipping.output['tip'])
tip.view(sim=tipping)

import numpy as np
import skfuzzy as fuzz
import matplotlib.pyplot as plt

x= np.arange(10,301,1)

y=np.arange(0,10)

from google.colab import files
uploaded = files.upload()

import pandas as pd
url = '1.01. Simple linear regression.csv'
dataframe = pd.read_csv(url)

import matplotlib.pyplot as plt

y=dataframe['GPA']

x=dataframe['SAT']

y

x

plt.scatter(x,y)
plt.xlabel('SAT')
plt.ylabel('GPA')
plt.show()

from google.colab import files
uploaded = files.upload()

import pandas as pd
url = 'mtcars (1).csv'
dataframe = pd.read_csv(url)

dataframe.hist()

pyplot.show()

correlations = dataframe.corr()
print(correlations)

y=dataframe['mpg']

x=dataframe['wt']

plt.scatter(x,y)
plt.xlabel('wt')
plt.ylabel('mpg')
plt.show()

dataframe.plot(kind='scatter',x='wt',y='mpg',figsize=(5,5),color='blue')
x=dataframe.drop(dataframe.columns[0:2],axis =1)

x=dataframe.drop(dataframe.columns[0:2],axis =1)

from sklearn import linear_model
#xac dinh mo hinh

regression_model = linear_model.LinearRegression()

from sklearn import linear_model
import pandas as pd

regression_model = linear_model.LinearRegression()
regression_model.fit(X = pd.DataFrame(dataframe['wt']),y = dataframe['mpg'])
print(regression_model.intercept_)
print(regression_model.coef_)

from sklearn import linear_model
import pandas as pd

#xac dinh mo hinh
x=dataframe['wt']
y=dataframe['mpg']
regression_model= linear_model.LinearRegression()

#openlab@hcmute.edu.vn

#gpa